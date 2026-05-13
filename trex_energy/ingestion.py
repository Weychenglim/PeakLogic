from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Iterable

from openpyxl import load_workbook
import pandas as pd


CANONICAL_COLUMNS = [
    "site_id",
    "interval_start",
    "interval_end",
    "kw_import",
    "kw_export",
    "kvar_import",
    "kvar_export",
    "has_solar",
    "existing_pv_kwp",
    "source_file",
    "source_sheet",
    "is_imputed",
]

ACTIVE_POWER_UNITS = {"auto", "kw", "kwh_per_interval"}


@dataclass(frozen=True)
class SiteMetadata:
    site_id: str
    has_solar: bool
    existing_pv_kwp: float | None
    source_file: str


def _infer_solar_flag(path: Path) -> bool:
    return "with solar" in path.stem.lower()


def _parse_capacity_candidate(value: object) -> float | None:
    if value is None:
        return None
    match = re.search(r"(\d+(?:\.\d+)?)\s*kWp", str(value), flags=re.IGNORECASE)
    if match:
        return float(match.group(1))
    return None


def _extract_existing_pv_kwp(workbook) -> float | None:
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(min_row=1, max_row=3, values_only=True):
            for value in row:
                parsed = _parse_capacity_candidate(value)
                if parsed is not None:
                    return parsed
    return None


def _row_values(values: Iterable[object], size: int = 6) -> list[object]:
    row = list(values)
    if len(row) < size:
        row.extend([None] * (size - len(row)))
    return row


def _active_unit_from_label(label: object) -> str | None:
    normalized = str(label or "").strip().lower().replace(" ", "_")
    if "kwh" in normalized:
        return "kwh_per_interval"
    if "kw" in normalized:
        return "kw"
    return None


def _resolve_active_unit(labels: Iterable[object], active_power_unit: str) -> str:
    if active_power_unit not in ACTIVE_POWER_UNITS:
        raise ValueError(f"active_power_unit must be one of {sorted(ACTIVE_POWER_UNITS)}")
    if active_power_unit != "auto":
        return active_power_unit

    for label in labels:
        inferred = _active_unit_from_label(label)
        if inferred is not None:
            return inferred
    return "kw"


def _to_kw(value: object, interval_start: pd.Timestamp, interval_end: pd.Timestamp, active_unit: str) -> float:
    numeric = float(value or 0.0)
    if active_unit == "kw":
        return numeric

    interval_hours = (interval_end - interval_start).total_seconds() / 3600.0
    if interval_hours <= 0:
        raise ValueError("Cannot convert interval energy to kW for a non-positive interval length")
    return numeric / interval_hours


def _extract_sheet_rows(
    path: Path,
    sheet_name: str,
    worksheet,
    active_power_unit: str = "auto",
) -> list[dict[str, object]]:
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    records: list[dict[str, object]] = []

    first_row = [str(value).strip().lower() if value is not None else "" for value in rows[0][:6]]
    first_row_normalized = [value.replace(" ", "_") for value in first_row]
    start_end_shape = first_row_normalized[:2] == ["start_time", "end_time"] and len(first_row_normalized) >= 6
    if start_end_shape and first_row_normalized[4:6] == ["kvar_export", "kvar_import"]:
        header_active_unit = _resolve_active_unit([rows[0][2], rows[0][3]], active_power_unit)
        if first_row_normalized[:6] == [
            "start_time",
            "end_time",
            "kw_export",
            "kw_import",
            "kvar_export",
            "kvar_import",
        ] or first_row_normalized[:6] == [
            "start_time",
            "end_time",
            "kwh_export",
            "kwh_import",
            "kvar_export",
            "kvar_import",
        ]:
            for raw in rows[1:]:
                start_time, end_time, kw_export, kw_import, kvar_export, kvar_import = _row_values(raw)
                if end_time is None:
                    continue
                interval_start = pd.Timestamp(start_time)
                interval_end = pd.Timestamp(end_time)
                records.append(
                    {
                        "interval_start": interval_start,
                        "interval_end": interval_end,
                        "kw_import": _to_kw(kw_import, interval_start, interval_end, header_active_unit),
                        "kw_export": _to_kw(kw_export, interval_start, interval_end, header_active_unit),
                        "kvar_import": float(kvar_import or 0.0),
                        "kvar_export": float(kvar_export or 0.0),
                        "source_sheet": sheet_name,
                    }
                )
            return records

    header_index = None
    for index, row in enumerate(rows[:6]):
        first_cell = str(row[0]).strip() if row and row[0] is not None else ""
        if first_cell == "Date / End Time":
            header_index = index
            break

    if header_index is None:
        return []

    header_row = rows[header_index]
    header_active_unit = _resolve_active_unit(_row_values(header_row, size=5)[1:3], active_power_unit)

    for raw in rows[header_index + 1 :]:
        timestamp, kw_export, kw_import, kvar_export, kvar_import = _row_values(raw, size=5)[:5]
        if timestamp is None:
            continue
        interval_end = pd.Timestamp(timestamp)
        interval_start = interval_end - pd.Timedelta(minutes=30)
        records.append(
            {
                "interval_start": interval_start,
                "interval_end": interval_end,
                "kw_import": _to_kw(kw_import, interval_start, interval_end, header_active_unit),
                "kw_export": _to_kw(kw_export, interval_start, interval_end, header_active_unit),
                "kvar_import": float(kvar_import or 0.0),
                "kvar_export": float(kvar_export or 0.0),
                "source_sheet": sheet_name,
            }
        )

    return records


def load_site_workbook(path: str | Path, active_power_unit: str = "auto") -> tuple[pd.DataFrame, SiteMetadata]:
    workbook_path = Path(path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        existing_pv_kwp = _extract_existing_pv_kwp(workbook)
        has_solar = _infer_solar_flag(workbook_path) or existing_pv_kwp is not None
        site_id = workbook_path.stem

        records: list[dict[str, object]] = []
        for sheet_name in workbook.sheetnames:
            records.extend(_extract_sheet_rows(workbook_path, sheet_name, workbook[sheet_name], active_power_unit))
    finally:
        workbook.close()

    frame = pd.DataFrame.from_records(records)
    if frame.empty:
        frame = pd.DataFrame(columns=CANONICAL_COLUMNS)
    else:
        frame = frame.sort_values("interval_end").reset_index(drop=True)
        frame["site_id"] = site_id
        frame["has_solar"] = has_solar
        frame["existing_pv_kwp"] = existing_pv_kwp
        frame["source_file"] = workbook_path.name
        frame["is_imputed"] = False
        frame = frame[CANONICAL_COLUMNS]

    metadata = SiteMetadata(
        site_id=site_id,
        has_solar=has_solar,
        existing_pv_kwp=existing_pv_kwp,
        source_file=workbook_path.name,
    )
    return frame, metadata
